# Full path for run script: "C:\Users\WINDOWS\miniconda3\envs\rat_lab\python.exe" -u "d:\DesktopApp_project\main.py"

import os
import sys
from turtle import pos


def app_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def resource_path(*parts):
    if getattr(sys, 'frozen', False):
        # When running from PyInstaller bundle
        # Check if using temporary extraction directory (sys._MEIPASS)
        if hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, *parts)

os.environ["TF_ENABLE_ONEDNN_OPTS"] = "1"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "2"
os.environ["OPENCV_SKIP_PYTHON_LOADER"] = "1"

# import warnings
# warnings.filterwarnings("ignore", category=DeprecationWarning)

import csv
import re

import matplotlib

matplotlib.use('Agg')

import cv2
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import tensorflow as tf
from matplotlib.backends.backend_qt5agg import \
    FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtCore import QDateTime, QLocale, QSize, Qt, QTimer
from PyQt5.QtGui import QColor, QFont, QIcon, QImage, QPainter, QPixmap
from PyQt5.QtWidgets import (QApplication, QColorDialog, QComboBox,
                             QDesktopWidget, QDialog, QDialogButtonBox,
                             QFileDialog, QFormLayout, QFrame, QGridLayout,
                             QHBoxLayout, QHeaderView, QInputDialog, QLabel,
                             QLineEdit, QMainWindow, QMessageBox, QPushButton,
                             QStackedWidget, QTableWidget, QTableWidgetItem,
                             QTextEdit, QVBoxLayout, QWidget)

from db import DatabaseManager


class IPCameraDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("กรอก IP ของกล้องหรือ Webcam")
        self.setFixedSize(500, 120)
        self.ip_input = QLineEdit(self)
        self.ip_input.setPlaceholderText("เช่น rtsp://admin:pass@192.168.1.64/stream1")

        buttons = QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        self.button_box = QDialogButtonBox(buttons)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        self.button_box.button(QDialogButtonBox.Ok).setObjectName("GreenButton")
        self.button_box.button(QDialogButtonBox.Cancel).setObjectName("RedButton")

        layout = QFormLayout()
        layout.addRow("IP/URL กล้องหรือ Webcam:", self.ip_input)
        layout.addWidget(self.button_box)
        self.setLayout(layout)

    def get_ip(self):
        return self.ip_input.text()



class PageOne(QWidget):
    def __init__(self, stack, main_window):
        super().__init__()
        self.stack = stack
        self.main_window = main_window
        self.video_path = None
        self.cap = None  
        self.is_playing = False
        self.setFixedSize(630, 670)  

        # Header
        self.header = QLabel("Mice Detection Program")
        self.header.setAlignment(Qt.AlignCenter)
        self.header.setFont(QFont("Arial", 26, QFont.Bold))
        self.header.setObjectName("Header")

        # Video / Camera Display
        self.label = QLabel("เลือกไฟล์วิดีโอหรือกล้อง")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setFixedSize(600, 400)
        self.label.setObjectName("VideoDisplay")
        self.label.mousePressEvent = self.on_label_click

        # Status Label
        self.status_label = QLabel("⏳ รอการเลือกวิดีโอหรือกล้อง")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setObjectName("StatusLabel")

        # Buttons
        self.button = QPushButton("เลือกไฟล์วิดีโอ")
        self.button.clicked.connect(self.browse_video)
        self.button.setObjectName("YellowButton")

        self.camera = QPushButton("ตรวจจับด้วยกล้อง")
        self.camera.clicked.connect(self.use_camera)
        self.camera.setObjectName("YellowButton")

        self.submit = QPushButton("เริ่มทดลอง")
        self.submit.clicked.connect(self.show_experiment_setup)
        self.submit.setObjectName("GreenButton")

        self.clear = QPushButton("ล้างข้อมูล")
        self.clear.clicked.connect(self.clear_data)
        self.clear.setObjectName("RedButton")

        self.switch_page = QPushButton("ไปยังหน้าสรุปข้อมูล")
        self.switch_page.clicked.connect(self.switch_to_summary_page)
        self.switch_page.setObjectName("MainButton")

        self.theme_dropdown = QComboBox()
        self.theme_dropdown.setFixedSize(180, 32)
        self.theme_dropdown.setObjectName("ThemeDropdown")

        self.theme_dropdown.addItem("🌞 Light Theme", "light")
        self.theme_dropdown.addItem("🌜 Dark Theme", "dark")
        self.theme_dropdown.addItem("🌸 Pastel Theme", "pastel")
        self.theme_dropdown.addItem("🌈 Default Theme", "default")

        current_index = self.theme_dropdown.findData(self.main_window.get_theme())
        if current_index != -1:
            self.theme_dropdown.setCurrentIndex(current_index)

        self.theme_dropdown.currentIndexChanged.connect(self.change_theme)

        # Layouts
        header_layout = QHBoxLayout()
        header_layout.addWidget(self.header)
        header_layout.addStretch() 
        header_layout.addWidget(self.theme_dropdown)

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.button)
        btn_layout.addWidget(self.camera)

        action_layout = QHBoxLayout()
        action_layout.addWidget(self.submit)
        action_layout.addWidget(self.clear)
        action_layout.addWidget(self.switch_page)

        center_layout = QVBoxLayout()
        center_layout.addWidget(self.label)
        center_layout.addWidget(self.status_label)
        center_layout.addSpacing(10)
        center_layout.addLayout(btn_layout)
        center_layout.addSpacing(15)
        center_layout.addLayout(action_layout)
        center_layout.setAlignment(Qt.AlignCenter)

        main_layout = QVBoxLayout()
        main_layout.addLayout(header_layout)  
        main_layout.addLayout(center_layout)
        main_layout.addStretch()
        self.setLayout(main_layout)

        self.timer = QTimer()
        self.timer.timeout.connect(self.next_frame)
    
    def update_theme_dropdown(self):
        current_theme = self.main_window.get_theme()
        index = self.theme_dropdown.findData(current_theme)
        if index != -1:
            self.theme_dropdown.blockSignals(True)
            self.theme_dropdown.setCurrentIndex(index)
            self.theme_dropdown.blockSignals(False)

    def on_label_click(self, event):
        if self.cap and self.cap.isOpened():
            if self.is_playing:
                self.timer.stop()
                self.status_label.setText("⏸️ วิดีโอถูกหยุด")
            else:
                self.timer.start(30)
                self.status_label.setText("▶️วิดีโอกำลังเล่น")
            self.is_playing = not self.is_playing

    def browse_video(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "เลือกไฟล์วิดีโอ",
            "",
            "Video Files (*.mp4 *.avi *.mov *.mkv);;All Files (*)"
        )
        if file_path:
            valid_extensions = [".mp4", ".avi", ".mov", ".mkv"]
            ext = os.path.splitext(file_path)[1].lower()

            if ext not in valid_extensions:
                self.label.setText("❌ ไม่สามารถเปิดไฟล์วิดีโอนี้ได้")
                return

            self.start_capture(file_path)

    def use_camera(self):
        dialog = IPCameraDialog()
        if dialog.exec_() == QDialog.Accepted:
            ip = dialog.get_ip()
            try:
                source = int(ip) 
            except ValueError:
                source = ip  
            self.start_capture(source)

    def start_capture(self, source):
        if self.cap:
            self.cap.release()

        self.video_path = source
        self.cap = cv2.VideoCapture(source)
        if not self.cap.isOpened():
            self.label.setText("❌ ไม่สามารถเชื่อมต่อกล้องหรือเปิดวิดีโอได้")
            return

        self.is_playing = True
        self.timer.start(30)
        self.status_label.setText("▶️วิดีโอกำลังเล่น")
        self.fps = self.cap.get(cv2.CAP_PROP_FPS)
        if self.fps == 0 or np.isnan(self.fps):  
            self.fps = 30

    def next_frame(self):
        if self.cap and self.is_playing:
            ret, frame = self.cap.read()
            if not ret:
                self.timer.stop()
                self.cap.release()
                self.label.setText("✅ จบวิดีโอหรือสัญญาณกล้องขาดหาย")
                return

            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = frame.shape
            qimg = QImage(frame.data, w, h, ch * w, QImage.Format_RGB888)

            fps_text = f"FPS: {self.fps:.1f}"
            painter = QPainter(qimg)
            painter.setPen(QColor(0, 255, 128))  
            font = QFont("Segoe UI", 28, QFont.Bold)
            painter.setFont(font)
            painter.drawText(20, 70, fps_text)
            painter.end()

            margin = 10  
            size = self.label.size()
            scaled_size = QSize(size.width() - margin * 2, size.height() - margin * 2)
            
            pixmap = QPixmap.fromImage(qimg).scaled(
                scaled_size, Qt.IgnoreAspectRatio, Qt.SmoothTransformation
            )

            self.label.setPixmap(pixmap)

    def show_experiment_setup(self):
        if self.video_path is None:
            self.label.setText("❌ ยังไม่ได้เลือกไฟล์วิดีโอหรือกล้อง")
            return

        preset_data = {"video_path": self.video_path}
        dialog = ExperimentSetupDialog(self.main_window.db, self, preset_data=preset_data)
        if dialog.exec_() == QDialog.Accepted:
            experiment_type_id, name, date, detail, video_path = dialog.get_experiment_data()
            if experiment_type_id == 0:
                QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณาเลือกประเภทการทดลอง!")
                return
            experiment_id = self.main_window.db.add_experiment(experiment_type_id, name, date, detail if detail is not None else "ไม่มีรายละเอียดการทดลอง", video_path)
            if experiment_id:
                self.main_window.set_experiment_name(name)
                self.main_window.set_experiment_id(experiment_id)
                self.main_window.set_fps(self.fps)
                self.main_window.set_video_path(self.video_path)
                self.main_window.switch_to_page(1)
                self.clear_data()
            else:
                QMessageBox.warning(self, "ข้อผิดพลาด", "ไม่สามารถบันทึกการทดลองได้ กรุณาลองใหม่!")
        else:
            self.label.setText("🎥 เลือกไฟล์วิดีโอหรือกล้อง")
    
    def change_theme(self, index):
        theme = self.theme_dropdown.itemData(index)
        if theme:
            self.main_window.set_theme(theme)
            self.main_window.load_theme(theme)

    def clear_data(self):
        if self.cap:
            self.cap.release()
        self.timer.stop()
        self.label.setText("🎥 เลือกไฟล์วิดีโอหรือกล้อง")
        self.status_label.setText("⏳ รอการเลือกวิดีโอหรือกล้อง")
        self.video_path = None
        self.cap = None
        self.is_playing = False

    def switch_to_summary_page(self):
        self.clear_data()
        self.main_window.switch_to_page(2)



class ExperimentSetupDialog(QDialog):
    def __init__(self, db, parent=None, preset_data=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("ตั้งค่าการทดลอง")
        self.setModal(True)  
        self.setFixedSize(570, 500)

        self.name_input = QLineEdit(self)
        self.date_input = QLineEdit(self)
        self.date_input.setText(
            QLocale(QLocale.English).toString(
                QDateTime.currentDateTime(), "yyyy-MM-dd HH:mm:ss"
            )
        )
        self.video_path_input = QLineEdit(self)
        self.type_combo = QComboBox(self)
        self.detail = QTextEdit(self)
        self.detail.setViewportMargins(10, 10, 10, 10)
        self.detail.setLineWrapMode(QTextEdit.NoWrap)
        self.detail.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.detail.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.detail.setFixedHeight(120)

        experiment_types = self.db.get_experiment_types()
        self.type_combo.addItem("เลือกประเภทการทดลอง", 0)
        for type_id, type_name in experiment_types:
            self.type_combo.addItem(type_name, type_id)
        
        if preset_data:
            if "name" in preset_data:
                self.name_input.setText(preset_data["name"])
            if "detail" in preset_data:
                self.detail.setPlainText(preset_data["detail"])
            if "date" in preset_data:
                self.date_input.setText(preset_data["date"])
                self.date_input.setReadOnly(True)
            if "type_id" in preset_data:
                index = self.type_combo.findData(preset_data["type_id"])
                if index != -1:
                    self.type_combo.setCurrentIndex(index)
            if "video_path" in preset_data:
                self.video_path_input.setText(preset_data["video_path"])

        layout = QFormLayout()
        layout.addRow("ที่อยู่ของวิดีโอหรือกล้อง:", self.video_path_input)
        layout.addRow("วันที่:", self.date_input)
        layout.addRow("ชื่อการทดลอง:", self.name_input)
        layout.addRow("ประเภทการทดลอง:", self.type_combo)
        layout.addRow("รายละเอียดการทดลอง:", self.detail)

        button_layout = QHBoxLayout()
        ok_button = QPushButton("ยืนยัน")
        ok_button.setObjectName("GreenButton")
        cancel_button = QPushButton("ยกเลิก")
        cancel_button.setObjectName("RedButton")
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)

        ok_button.clicked.connect(self.validate_and_accept)
        cancel_button.clicked.connect(self.reject)

    def validate_and_accept(self):
        name = self.name_input.text().strip()
        experiment_type_id = self.type_combo.currentData()
        if not name:
            QMessageBox.warning(self, "ข้อมูลไม่ครบ", "กรุณากรอกชื่อการทดลองก่อนกด ยืนยัน")
            return
        
        elif experiment_type_id == 0:
            QMessageBox.warning(self, "ข้อมูลไม่ครบ", "กรุณาเลือกประเภทการทดลองก่อนกด ยืนยัน")
            return
        self.accept()

    def get_experiment_data(self):
        experiment_type_id = self.type_combo.currentData()
        name = self.name_input.text()
        date = self.date_input.text()
        video_path = self.video_path_input.text()
        detail = self.detail.toPlainText()
        if not name: 
            raise ValueError("ชื่อการทดลองต้องไม่ว่าง")
        return experiment_type_id, name, date, detail, video_path



class PageTwo(QWidget):
    def __init__(self, stack, main_window):
        super().__init__()
        self.stack = stack
        self.main_window = main_window
        self.setFocusPolicy(Qt.StrongFocus)
        self.move_mode = False
        self.is_camera = False
        self.setFixedSize(1400, 950)

        self.is_playing = False
        self.cap = None
        self.last_frame = None
        self.frame_count = 0
        self.process_every_n = 2
        self.hide_ui = False
        self.hide_detail_notes = True
        self.running_time = 0
        self.ex_timer = 0
        self.motion_step = 1.0
        self.motion_percent = 5.0
        self.last_binary_shape = None
        

        model_path = resource_path("model", "model_for_rat_V3.keras")
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"Model file not found: {model_path}")

        
        self.model = tf.keras.models.load_model(model_path, safe_mode=False)
        self.polygon_manager = PolygonManager()
        self.drawing_polygon = False
        self.active_started = False
        self.mouse_pos = None
        self.polygon_name = ""
        self.polygon_color = (0, 255, 0)
        self.raw_data = []

        self.video_label = QLabel("📹 เริ่มการตรวจตำแหน่ง")
        self.video_label.setAlignment(Qt.AlignCenter)
        self.video_label.setObjectName("VideoDisplay")
        self.video_label.mousePressEvent = self.polygon_draw

        self.create_polygon_table()

        self.back_button = QPushButton("ย้อนกลับ")
        self.back_button.clicked.connect(self.on_back_button_clicked)
        self.back_button.setObjectName("YellowButton")

        self.summary_button = QPushButton("สรุปผล")
        self.summary_button.clicked.connect(self.submit_summary)
        self.summary_button.setObjectName("GreenButton")

        self.detail_note = QPushButton("รายละเอียดการทดลอง")
        self.detail_note.clicked.connect(self.show_detail_notes)
        self.detail_note.setObjectName("MainButton")

        self.restart_button = QPushButton("เริ่มทดลอง/จับเวลาใหม่")
        self.restart_button.clicked.connect(self.restart_experiment)
        self.restart_button.setObjectName("GreenButton")

        self.experiment_timer = QPushButton("ตั้งเวลาการทดลอง")
        self.experiment_timer.clicked.connect(self.set_timer)
        self.experiment_timer.setObjectName("MainButton")

        button_layout = QVBoxLayout()
        button_layout.addWidget(self.restart_button)
        button_layout.addWidget(self.detail_note)
        button_layout.addWidget(self.experiment_timer)
        button_layout.addWidget(self.back_button)
        button_layout.addWidget(self.summary_button)
        

        table_layout = QHBoxLayout()
        table_layout.addWidget(self.polygon_table)
        table_layout.addLayout(button_layout)

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.video_label, stretch=1)
        main_layout.addLayout(table_layout)

        self.setLayout(main_layout)

        self.timer = QTimer()
        self.timer.timeout.connect(self.next_frame)
        self.setFocus()

    def restart_experiment(self):
        self.frame_count = 0
        self.running_time = 0
        self.raw_data = []

        for polygon in self.polygon_manager.polygons.values():
            polygon.hit_count = 0
            polygon.hit_time = 0
            polygon.is_inside = False
            polygon.prev_rat_mask = None

        self.update_polygon_table()
        if self.last_frame is not None:
            current_frame = self.last_frame.copy()

            margin = 10
            size = self.video_label.size()
            scaled_size = QSize(size.width() - margin * 2, size.height() - margin * 2)

            frame_display = cv2.resize(
                current_frame,
                (scaled_size.width(), scaled_size.height()),
                interpolation=cv2.INTER_CUBIC
            )

            self.polygon_manager.draw_all(frame_display)

            overlay_frame = cv2.cvtColor(frame_display, cv2.COLOR_BGR2RGB)
            h, w, ch = overlay_frame.shape
            qimg = QImage(overlay_frame.data, w, h, ch * w, QImage.Format_RGB888)

            painter = QPainter(qimg)
            match self.main_window.current_theme:
                case "dark":
                    painter.setPen(QColor(0, 255, 128))
                case "light":
                    painter.setPen(QColor(0, 128, 255))
                case "pastel":
                    painter.setPen(QColor(255, 0, 128))
                case "default":
                    painter.setPen(QColor(200, 200, 200))
                case _:
                    painter.setPen(QColor(200, 200, 200))

            font = QFont("Segoe UI", 14)
            painter.setFont(font)

            if not self.hide_ui:
                painter.drawText(10, 60, f"ID name: {self.experiment_name}")
                painter.drawText(10, 110, f"FPS: {self.fps:.1f}")
                painter.drawText(10, 140, f"Time: 0:00")
                painter.drawText(10, 170, f"(Space) Play/Pause")
                painter.drawText(10, 200, f"(Z) Draw Mode")
                painter.drawText(10, 230, f"(M) Move Mode")
                painter.drawText(10, 260, f"(N) Change Experiment ID")
                painter.drawText(10, 290, f"(H) Hide UI")
                painter.drawText(10, 320, f"(V) View Detail Notes")
                painter.drawText(10, 350, f"(T) Set Timer")
                painter.drawText(10, 380, f"(X) Clear All!")
                if self.is_motion_experiment():
                    painter.drawText(10, 410, f"(P) Set Motion % : {self.motion_percent:.1f}%")
                if self.move_mode:
                    painter.drawText(10, 420, f"Moving Polygon ...")
                    painter.drawText(10, 450, f"(W, A, S, D) to Move")
                    painter.drawText(10, 480, f"(Q, E) to Rotate")
                elif self.drawing_polygon:
                    painter.drawText(10, 420, f"Drawing Polygon ...")
                    painter.drawText(10, 450, f"(Left Click) to Draw")
                    painter.drawText(10, 480, f"(Right Click) to Close")

                if not self.hide_detail_notes:
                    painter.drawText(10, 520, f"Detail Notes:")
                    painter.drawText(10, 550, f"{self.experiment_note}")

            painter.end()

            pixmap = QPixmap.fromImage(qimg).scaled(
                scaled_size, Qt.IgnoreAspectRatio, Qt.SmoothTransformation
            )
            self.video_label.setPixmap(pixmap)
        self.setFocus()

    def set_timer(self):
        while True:
            get_timer, ok = QInputDialog.getText(self, "ตั้งเวลาการทดลอง", "กรุณาตั้งเวลาเป็นหน่วย \"นาที\"", text=str(self.ex_timer // 60))
            if not ok or not get_timer.strip():
                self.setFocus()
                return  
            try:
                self.ex_timer = int(get_timer) * 60
                break  
            except ValueError:
                QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณากรอกเวลาเป็นตัวเลขเท่านั้น")
        self.setFocus()
    
    def is_time_out(self):
        if self.ex_timer == 0:
            return 
        elif self.ex_timer == self.running_time:
            self.stop_video()
            self.submit_summary(skip_dialog=True)

    def keyPressEvent(self, event):
        key = event.key()
        if key == Qt.Key_Z and not self.drawing_polygon:
            self.start_drawing()
            self.next_frame()
        elif key == Qt.Key_Space:
            self.on_label_click()
        elif key == Qt.Key_H:
            self.hide_ui = not self.hide_ui
            self.next_frame()
        elif key == Qt.Key_M:
            if not self.drawing_polygon:
                self.move_polygon()
        elif key == Qt.Key_N:
            self.change_experiment_name()
        elif key == Qt.Key_X:
            self.clear_all_data()
            self.next_frame()
            self.update_polygon_table()
        elif key == Qt.Key_V:
            self.show_detail_notes()
            self.next_frame()
        elif key == Qt.Key_T:
            self.set_timer()
            self.next_frame()
        elif key == Qt.Key_P:
            self.set_motion_percent()
            self.next_frame()
        elif key == Qt.Key_BracketLeft:   
            self.motion_percent = max(0.0, self.motion_percent - self.motion_step)
            self.next_frame()
        elif key == Qt.Key_BracketRight:  
            self.motion_percent = min(100.0, self.motion_percent + self.motion_step)
            self.next_frame()
            
        elif self.move_mode:
            move_distance = 15
            if key == Qt.Key_W:
                self.polygon_manager.move_all_polygons(0, -move_distance)
            elif key == Qt.Key_S:
                self.polygon_manager.move_all_polygons(0, move_distance)
            elif key == Qt.Key_A:
                self.polygon_manager.move_all_polygons(-move_distance, 0)
            elif key == Qt.Key_D:
                self.polygon_manager.move_all_polygons(move_distance, 0)
            elif key == Qt.Key_Q:
                self.polygon_manager.rotate_all_polygons(np.deg2rad(5))
            elif key == Qt.Key_E:
                self.polygon_manager.rotate_all_polygons(np.deg2rad(-5))
            self.next_frame()
        else:
            super().keyPressEvent(event)
    def set_motion_percent(self):
        text, ok = QInputDialog.getText(
            self, "ตั้งค่า % การขยับ", "กำหนด % การขยับ (0-100):",
            text=str(self.motion_percent)
        )
        if not ok or not text.strip():
            self.setFocus()
            return
        try:
            v = float(text)
            if v < 0 or v > 100:
                QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณากรอกค่า 0-100 เท่านั้น")
                self.setFocus()
                return
            self.motion_percent = v
        except ValueError:
            QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณากรอกเป็นตัวเลขเท่านั้น")
        self.setFocus()
        
    def show_detail_notes(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("รายละเอียดการทดลอง")
        dialog.resize(550, 400)  
        layout = QVBoxLayout(dialog)

        label = QLabel("สามารถแก้ไขรายละเอียดการทดลองได้")
        layout.addWidget(label)

        text_edit = QTextEdit()
        text_edit.setPlainText(self.experiment_note) 
        text_edit.setViewportMargins(10, 10, 10, 10)
        text_edit.setLineWrapMode(QTextEdit.NoWrap)  
        text_edit.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)  
        text_edit.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        layout.addWidget(text_edit)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec_() == QDialog.Accepted:
            new_detail = text_edit.toPlainText().strip()
            if new_detail:
                experiment_id = self.experiment_id
                self.main_window.db.update_experiment_detail_note(experiment_id, new_detail)
                self.experiment_note = new_detail
        self.setFocus()

    def change_experiment_name(self):
        self.experiment_name, ok = QInputDialog.getText(self, "แก้ไขรหัสการทดลอง", "กรุณาตั้งรหัส", text=self.experiment_name)
        if not ok or not self.experiment_name.strip():
            self.setFocus()
            return
        self.setFocus()
        experiment_id = self.experiment_id  
        if self.main_window.db.update_experiment_name(experiment_id, self.experiment_name):
            return
        
    def clear_all_data(self):
        self.polygon_manager.polygons = {}
        self.main_window.db.delete_area_summary_by_experiment_id(self.experiment_id)

    def move_polygon(self):
        self.move_mode = not self.move_mode

    def delete_polygon(self, name):
        if name in self.polygon_manager.polygons:
            area_id = self.polygon_manager.polygons[name].id
            del self.polygon_manager.polygons[name]

            if area_id is not None:
                self.main_window.db.delete_area_summary_by_id(area_id)
            self.drawing_polygon = False
            self.active_started = False
            self.polygon_manager.active_polygon = None

            self.update_polygon_table()
            self.next_frame()
            self.setFocus()

    def edit_polygon(self, name):
        polygon = self.polygon_manager.polygons.get(name)
        if not polygon:
            return

        new_name, ok = QInputDialog.getText(self, "แก้ไขชื่อพื้นที่", "กรุณาตั้งชื่อใหม่", text=polygon.name)
        if not ok or not new_name.strip():
            return

        new_color = QColorDialog.getColor(QColor(polygon.color[2], polygon.color[1], polygon.color[0]), self)
        if not new_color.isValid():
            return

        if new_name != name:
            self.polygon_manager.polygons[new_name] = self.polygon_manager.polygons.pop(name)
            polygon.name = new_name

        polygon.color = (new_color.blue(), new_color.green(), new_color.red())

    def update_polygon_table(self):
        polygons = self.polygon_manager.polygons
        self.polygon_table.setRowCount(len(polygons))

        for i, (name, polygon) in enumerate(polygons.items()):
            name_item = QTableWidgetItem(name)
            color_item = QTableWidgetItem()
            hit_count_item = QTableWidgetItem(str(polygon.hit_count))
            hit_time_item = QTableWidgetItem(str(round(polygon.hit_time, 2)) + " วินาที")

            color = polygon.color
            rgb_color = (color[2], color[1], color[0])
            qcolor = QColor(*rgb_color)
            color_item.setBackground(qcolor)
            color_item.setText(f"({color[2]}, {color[1]}, {color[0]})")

            btn_delete = QPushButton("🗑️ ลบ")
            btn_delete.setStyleSheet("background-color: #DC2626; color: #000000;")
            btn_delete.clicked.connect(lambda _, name=name: self.delete_polygon(name))

            btn_edit = QPushButton("✏️แก้ไข")
            btn_edit.setStyleSheet("background-color: #F59E0B; color: #000000;")
            btn_edit.clicked.connect(lambda _, name=name: self.edit_polygon(name))

            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.addWidget(btn_delete)
            layout.addWidget(btn_edit)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setAlignment(Qt.AlignCenter)
            cell_widget.setLayout(layout)

            self.polygon_table.setItem(i, 0, name_item)
            self.polygon_table.setItem(i, 1, color_item)
            self.polygon_table.setItem(i, 2, hit_count_item)
            self.polygon_table.setItem(i, 3, hit_time_item)
            self.polygon_table.setCellWidget(i, 4, cell_widget)

    def create_polygon_table(self):
        self.polygon_table = QTableWidget()
        self.polygon_table.setColumnCount(5)
        self.polygon_table.setHorizontalHeaderLabels(["ชื่อ Polygon", "สี", "จำนวนครั้ง", "เวลาทั้งหมด", "การจัดการ"])
        
        self.polygon_table.verticalHeader().setDefaultSectionSize(50)

        header = self.polygon_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)

        self.polygon_table.setFixedHeight(220)
        self.polygon_table.setEditTriggers(QTableWidget.NoEditTriggers)

        font = self.polygon_table.font()
        font.setPointSize(8)
        self.polygon_table.setFont(font)

        self.update_polygon_table()

    def stop_video(self):
        self.timer.stop()
        if self.cap:
            self.cap.release()
            self.cap = None
        self.is_playing = False

    def on_back_button_clicked(self):
        self.stop_video()
        self.clear_all_data()
        self.main_window.db.delete_area_summary_by_experiment_id(self.experiment_id)
        self.main_window.db.delete_experiment_by_id(self.experiment_id)
        self.main_window.set_experiment_id(None)
        self.main_window.switch_to_page(0)

    def is_ip_camera(self):
        path = self.video_path
        if isinstance(path, int):
            return True
        elif isinstance(path, str):
            return (path.startswith("rtsp://") or 
                    path.startswith("http://") or 
                    path.startswith("https://") or 
                    path.startswith("/dev/video"))
        else:
            return False

    def update_video(self):
        self.stop_video()
        self.polygon_manager.polygons = {}
        self.polygon_manager.active_polygon = None
        self.drawing_polygon = False
        self.active_started = False
        self.raw_data = []
        self.frame_count = 0
        self.running_time = 0
        self.last_frame = None
        self.last_binary_shape = None

        self.update_polygon_table()

        self.experiment_name = self.main_window.get_experiment_name()
        self.experiment_id = self.main_window.get_experiment_id()
        self.experiment_type_id = self.main_window.db.get_experiment_type_id_by_experiment_id(self.experiment_id)
        self.experiment_type_name = self.main_window.db.get_experiment_type_by_id(self.experiment_type_id)
        self.experiment_note = self.main_window.db.get_experiment_note_by_id(self.experiment_id)
        self.fps = self.main_window.get_fps()
        self.video_path = self.main_window.get_video_path()

        if self.video_path is not None:
            if self.cap:
                self.cap.release()
            self.cap = cv2.VideoCapture(self.video_path)
            if not self.cap.isOpened():
                print("Error: ไม่สามารถเปิดวิดีโอจาก Path นี้ได้")
                self.cap = None
                return

            self.is_camera = self.is_ip_camera()

            if self.is_camera:
                self.is_playing = True
                self.timer.start(30)
            else:
                self.is_playing = True
                self.next_frame()
                self.timer.start(30)
        else:
            print("Warning: video_path is None")

        self.setFocus()
            
    def is_motion_experiment(self) -> bool:
        name = (self.experiment_type_name or "").lower()
        return ("tail suspension" in name) or ("tst" in name) or ("forced swim" in name) or ("fst" in name)
    
    def on_label_click(self):
        if self.cap and self.cap.isOpened():
            if self.is_playing:
                self.timer.stop()
            else:
                self.timer.start(30)
            self.is_playing = not self.is_playing

    def next_frame(self):
        if not self.cap:
            return

        if not self.is_playing:
            if self.last_frame is None:
                success, frame = self.cap.read()
                if not success:
                    return
                self.last_frame = frame
            current_frame = self.last_frame
        else:
            ret, frame = self.cap.read()
            if not ret:
                self.timer.stop()
                self.cap.release()
                self.is_playing = False
                return
            self.last_frame = frame
            self.frame_count += 1
            current_frame = frame

        self.is_time_out()

        margin = 10
        size = self.video_label.size()
        scaled_size = QSize(size.width() - margin * 2, size.height() - margin * 2)
        frame_display = cv2.resize(
            current_frame,
            (scaled_size.width(), scaled_size.height()),
            interpolation=cv2.INTER_CUBIC
        )

        input_frame = cv2.resize(current_frame, (128, 128), interpolation=cv2.INTER_CUBIC)
        input_frame = np.expand_dims(input_frame, axis=0)

        try:
            result = self.model.predict(input_frame, verbose=0)[0]
        except Exception as e:
            self.timer.stop()
            self.is_playing = False
            self.video_label.setText(f"❌ Model predict error: {e}")
            return

        result = np.squeeze(result)
        result = np.clip(result, 0, 1)

        mask = cv2.resize(
            result,
            (scaled_size.width(), scaled_size.height()),
            interpolation=cv2.INTER_LINEAR
        )
        mask = (mask * 255).astype(np.uint8)

        mask = cv2.GaussianBlur(mask, (5, 5), 0)
        _, binary_mask = cv2.threshold(mask, 110, 255, cv2.THRESH_BINARY)

        kernel = np.ones((3, 3), np.uint8)
        binary_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_OPEN, kernel, iterations=1)
        binary_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_CLOSE, kernel, iterations=2)

        self.last_binary_shape = binary_mask.shape[:2]

        contours, _ = cv2.findContours(binary_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        centers = []

        if contours:
            # TST/FST process 2 contours, other experiments process 1
            max_contours = 2 if self.is_motion_experiment() else 1
            contours = sorted(contours, key=cv2.contourArea, reverse=True)[:max_contours]
            for cnt in contours:
                M = cv2.moments(cnt)
                if M["m00"] != 0:
                    cX = int(M["m10"] / M["m00"])
                    cY = int(M["m01"] / M["m00"])
                    centers.append((cX, cY))
                    cv2.circle(frame_display, (cX, cY), 5, (0, 0, 255), -1)

        green_layer = np.zeros((binary_mask.shape[0], binary_mask.shape[1], 3), dtype=np.uint8)
        green_layer[:, :] = (0, 255, 0)
        green_masked = cv2.bitwise_and(green_layer, green_layer, mask=binary_mask)

        frame_display = frame_display.astype(np.uint8)
        overlay_frame = cv2.addWeighted(frame_display, 0.9, green_masked, 0.5, 0)

        self.polygon_manager.draw_all(overlay_frame)

        if self.is_motion_experiment():
            self.calculate_motion_time(contours)
        else:
            self.calculate_time_in_area(contours)

        if self.fps and self.fps > 0 and self.frame_count % max(1, int(self.fps)) == 0:
            self.update_polygon_table()

        overlay_frame = cv2.cvtColor(overlay_frame, cv2.COLOR_BGR2RGB)
        h, w, ch = overlay_frame.shape
        qimg = QImage(overlay_frame.data, w, h, ch * w, QImage.Format_RGB888)

        painter = QPainter(qimg)
        match self.main_window.current_theme:
            case "dark":
                painter.setPen(QColor(0, 255, 128))
            case "light":
                painter.setPen(QColor(0, 128, 255))
            case "pastel":
                painter.setPen(QColor(255, 0, 128))
            case "default":
                painter.setPen(QColor(200, 200, 200))
            case _:
                painter.setPen(QColor(200, 200, 200))

        font = QFont("Segoe UI", 14)
        painter.setFont(font)
        format_time = f"Time: {self.format_time()}"

        if not self.hide_ui:
            painter.drawText(10, 60, f"ID name: {self.experiment_name}")
            painter.drawText(10, 110, f"FPS: {self.fps:.1f}")
            painter.drawText(10, 140, format_time)
            painter.drawText(10, 170, f"(Space) Play/Pause")
            painter.drawText(10, 200, f"(Z) Draw Mode")
            painter.drawText(10, 230, f"(M) Move Mode")
            painter.drawText(10, 260, f"(N) Change Experiment ID")
            painter.drawText(10, 290, f"(H) Hide UI")
            painter.drawText(10, 320, f"(V) View Detail Notes")
            painter.drawText(10, 350, f"(T) Set Timer")
            painter.drawText(10, 380, f"(X) Clear All!")
            if self.is_motion_experiment():
                painter.drawText(10, 410, f"(P) Set Motion % : {self.motion_percent:.1f}%")
            if self.move_mode:
                painter.drawText(10, 470, f"Moving Polygon ...")
                painter.drawText(10, 500, f"(W, A, S, D) to Move")
                painter.drawText(10, 530, f"(Q, E) to Rotate")
            elif self.drawing_polygon:
                painter.drawText(10, 470, f"Drawing Polygon ...")
                painter.drawText(10, 500, f"(Left Click) to Draw")
                painter.drawText(10, 530, f"(Right Click) to Close")

            if not self.hide_detail_notes:
                painter.drawText(10, 520, f"Detail Notes:")
                painter.drawText(10, 550, f"{self.experiment_note}")

        painter.end()

        pixmap = QPixmap.fromImage(qimg).scaled(
            scaled_size, Qt.IgnoreAspectRatio, Qt.SmoothTransformation
        )
        self.video_label.setPixmap(pixmap)
        
    def format_time(self):
        seconds = int(self.frame_count // self.fps)
        minutes = seconds // 60  
        remaining_seconds = seconds % 60  
        self.running_time = seconds
        return f"{minutes}:{remaining_seconds:02d}"

    def start_drawing(self):
        existing_names = [polygon.name for polygon in self.polygon_manager.polygons.values()]
        while True:
            name, ok = QInputDialog.getText(self, "ชื่อพื้นที่", "กรุณาตั้งชื่อ")
            if not ok:
                return

            name = name.strip()
            if not name:
                base_name = f"Polygon-{len(self.polygon_manager.polygons) + 1}"
                name = base_name
                i = 2
                while name in existing_names:
                    name = f"{base_name}-{i}"
                    i += 1
                break

            if name in existing_names:
                QMessageBox.warning(self, "ชื่อซ้ำ", f'ชื่อ "{name}" ถูกใช้ไปแล้ว กรุณาตั้งชื่อใหม่')
            else:
                break

        color = QColorDialog.getColor()
        if not color.isValid():
            self.setFocus()
            return

        self.drawing_polygon = True
        self.active_started = False
        self.polygon_name = name
        self.polygon_color = (color.blue(), color.green(), color.red())
        self.setFocus()
    def polygon_draw(self, event):
        pos = self.video_label.mapFrom(self, event.pos())
        x = pos.x()
        y = pos.y()
        if event.button() == Qt.RightButton and self.drawing_polygon:
            self.polygon_manager.close_active()
            self.drawing_polygon = False
            self.active_started = False
            self.polygon_name = ""
            self.polygon_color = (0, 255, 0)

            if self.polygon_manager.active_polygon:
                polygon = self.polygon_manager.active_polygon
                experiment_id = self.experiment_id
                area_points = str(polygon.points)
                color_str = str(polygon.color)

                polygon_id = self.main_window.db.save_area_summary(
                    experiment_id,
                    polygon.name,
                    color_str,
                    polygon.hit_count,
                    polygon.hit_time,
                    area_points
                )
                if polygon_id is not None:
                    print(f"Polygon saved with ID: {polygon_id} Name : {polygon.name}")
                    polygon.id = polygon_id

            self.update_polygon_table()
            self.next_frame()
            self.setFocus()
            return

        if event.button() == Qt.LeftButton:
            if self.drawing_polygon:
                if not self.active_started:
                    name = self.polygon_name
                    color = self.polygon_color
                    self.polygon_manager.new_polygon(name, color)
                    self.active_started = True

                self.polygon_manager.add_point_to_active((x, y))
                self.update_polygon_table()
                self.next_frame()
                self.setFocus()

    def calculate_time_in_area(self, contours):
        experiment_id = self.experiment_id
        fps = self.fps if self.fps and self.fps > 0 else 30
        dt = 1 / fps

        time_stamp = round(self.frame_count / fps, 2)

    
        current_occupancy = {}  
        rat_positions = [] 
        
        for polygon in self.polygon_manager.polygons.values():
            current_occupancy[id(polygon)] = False


        for cnt in contours[:1]:
            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cX = int(M["m10"] / M["m00"])
            cY = int(M["m01"] / M["m00"])

            best_poly = None
            for polygon in self.polygon_manager.polygons.values():
                if not polygon.is_closed:
                    continue
                pts = np.array(polygon.points, dtype=np.int32)
                if cv2.pointPolygonTest(pts, (float(cX), float(cY)), False) >= 0:
                    best_poly = polygon
                    break

            rat_positions.append((cX, cY, best_poly))
            if best_poly is not None:
                current_occupancy[id(best_poly)] = True

        # Second pass: process rats and update hit counts
        for cX, cY, best_poly in rat_positions:
            # Detect entering the area (transition from outside to inside)
            if best_poly is not None:
                if not best_poly.is_inside:
                    best_poly.hit_count += 1
                best_poly.is_inside = True
                best_poly.hit_time += dt

            self.raw_data.append({
                'experiment_id': experiment_id,
                'area_id': best_poly.id if best_poly else None,
                'time_stamp': time_stamp,
                'frame_count': self.frame_count,
                'area_name': best_poly.name if best_poly else None,
                'rat_position_x': cX,
                'rat_position_y': cY,
            })
        
        for polygon in self.polygon_manager.polygons.values():
            if not current_occupancy[id(polygon)]:
                polygon.is_inside = False
    def calculate_motion_time(self, contours):
        experiment_id = self.experiment_id
        fps = self.fps if self.fps and self.fps > 0 else 30
        dt = 1 / fps
        h, w = self.last_binary_shape
        time_stamp = round(self.frame_count / fps, 2)

        for cnt in contours[:2]:
            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue

            cX = int(M["m10"] / M["m00"])
            cY = int(M["m01"] / M["m00"])

            rat_mask = np.zeros((h, w), dtype=np.uint8)
            cv2.drawContours(rat_mask, [cnt], -1, 255, thickness=-1)
            rat_area = int(np.count_nonzero(rat_mask))
            if rat_area == 0:
                continue
            best_poly = None
            for polygon in self.polygon_manager.polygons.values():
                if not polygon.is_closed:
                    continue
                pts = np.array(polygon.points, dtype=np.int32)
                if cv2.pointPolygonTest(pts, (float(cX), float(cY)), False) >= 0:
                    best_poly = polygon
                    break
            moving = False
            motion_change = 0.0

            if best_poly is not None:
                prev = best_poly.prev_rat_mask
                if prev is not None:
                    xor_area = int(np.count_nonzero(cv2.bitwise_xor(prev, rat_mask)))
                    motion_change = (xor_area / rat_area) * 100.0
                    moving = motion_change >= self.motion_percent
                else:
                    moving = True

                best_poly.prev_rat_mask = rat_mask
                best_poly.is_inside = moving

                if not moving:
                    best_poly.hit_time += dt
            else:
                for polygon in self.polygon_manager.polygons.values():
                    polygon.is_inside = False

            self.raw_data.append({
                'experiment_id': experiment_id,
                'area_id': best_poly.id if best_poly else None,
                'time_stamp': time_stamp,
                'frame_count': self.frame_count,
                'area_name': best_poly.name if best_poly else None,
                'rat_position_x': cX,
                'rat_position_y': cY,
            })
        
    def submit_summary(self, skip_dialog=False):
        if not skip_dialog:
            reply = QMessageBox.question(self, "ยืนยันการจบการทดลอง", 
                                        "คุณต้องการจบการทดลองและบันทึกข้อมูลหรือไม่?",
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return  
            
        for polygon in self.polygon_manager.polygons.values():
            if polygon.id is not None:
                success = self.main_window.db.update_area_summary(polygon.id, polygon.name, str(polygon.color), polygon.hit_count, round(polygon.hit_time, 2), str(polygon.points))
                if not success:
                    print(f"ไม่สามารถอัปเดตข้อมูล polygon {polygon.name} (ID: {polygon.id})")

        if self.raw_data:
            all_area_id = [polygon.id for polygon in self.polygon_manager.polygons.values() if polygon.id is not None]
            all_area_name = [polygon.name for polygon in self.polygon_manager.polygons.values() if polygon.id is not None]
            data_to_insert = [
                (d['experiment_id'], 
                 d['area_id'] if d['area_id'] in all_area_id else None, 
                 d['time_stamp'], 
                 d['frame_count'], 
                 d['area_name'] if d['area_name'] in all_area_name else None, 
                 d['rat_position_x'], 
                 d['rat_position_y'])
                for d in self.raw_data
            ]


            success = self.main_window.db.save_raw_data_batch(data_to_insert)
            if not success:
                QMessageBox.warning(self, "ข้อผิดพลาด", "ไม่สามารถบันทึก raw data ได้")
            self.raw_data = []

        self.stop_video()
        self.clear_all_data()
        self.main_window.switch_to_page(2)



class Polygon:
    def __init__(self, name, color):
        self.id = None
        self.points = []
        self.color = color
        self.name = name
        self.is_closed = False
        self.hit_count = 0
        self.hit_time = 0
        self.is_inside = False
        self.prev_rat_mask = None
        
    
    def add_point(self, point):
        if not self.is_closed:
            self.points.append(point)
    
    def close(self):
        if len(self.points) >= 3:
            self.is_closed = True
    
    def draw(self, frame, thickness=2):
        if len(self.points) >= 2:
            pts = np.array(self.points, dtype=np.int32)
            cv2.polylines(frame, [pts], isClosed=self.is_closed, color=self.color, thickness=thickness)
        for x, y in self.points:
            cv2.circle(frame, (int(x), int(y)), 3, (255, 0, 0), -1)
        if self.points:
            cv2.putText(frame, self.name, (int(self.points[0][0]), int(self.points[0][1])), cv2.FONT_HERSHEY_SIMPLEX, 0.6, self.color, 2)



class PolygonManager:
    def __init__(self):
        self.polygons = {}
        self.active_polygon = None

    def new_polygon(self, name, color):
        self.polygons[name] = Polygon(name, color)
        self.active_polygon = self.polygons[name]

    def add_point_to_active(self, point):
        if self.active_polygon:
            self.active_polygon.add_point(point)

    def close_active(self):
        if self.active_polygon:
            self.active_polygon.close()

    def draw_all(self, frame):
        for polygon in self.polygons.values():
            polygon.draw(frame)

    def move_all_polygons(self, dx, dy):
        for polygon in self.polygons.values():
            if polygon.is_closed:
                polygon.points = [(x + dx, y + dy) for x, y in polygon.points]
    
    def rotate_all_polygons(self, theta):
        centers = []
        for polygon in self.polygons.values():
            if polygon.is_closed and polygon.points:
                pts = np.array(polygon.points, dtype=np.float32)
                cx, cy = np.mean(pts, axis=0)
                centers.append((cx, cy))

        if not centers:
            return

        global_cx = np.mean([c[0] for c in centers])
        global_cy = np.mean([c[1] for c in centers])

        cos_theta = np.cos(theta)
        sin_theta = np.sin(theta)

        for polygon in self.polygons.values():
            if polygon.is_closed and polygon.points:
                new_points = []
                for x, y in polygon.points:
                    x_shifted = x - global_cx
                    y_shifted = y - global_cy

                    x_rot = x_shifted * cos_theta - y_shifted * sin_theta
                    y_rot = x_shifted * sin_theta + y_shifted * cos_theta

                    x_new = x_rot + global_cx
                    y_new = y_rot + global_cy
                    new_points.append((x_new, y_new))
            polygon.points = new_points



class PageThree(QWidget):
    def __init__(self, stack, main_window):
        super().__init__()
        self.stack = stack
        self.main_window = main_window
        self.setFixedSize(1400, 950)
        self.area_summary = None
        self.raw_data = None
        self.df_raw_data = None
        self.df_area_summary = None
        self.current_experiment_id = None
        self.current_experiment_name = None
        self.current_experiment_date = None
        self.current_experiment_type = None
        self.current_experiment_note = None
        self.total_time = "ไม่มีปรากฏข้อมูล"
        self.total_area = "ไม่มีปรากฎข้อมูล"
        self.experiment_time = "ไม่มีปรากฏข้อมูล"
        self.avg_time = "ไม่มีปรากฏข้อมูล"
        self.existing_data = None

        self.switch_page = QPushButton("กลับไปยังหน้าแรก")
        self.switch_page.clicked.connect(self.switch_to_home_page)
        self.switch_page.setObjectName("MainButton")

        self.csv_export = QPushButton("Export to CSV")
        self.csv_export.clicked.connect(self.export_to_csv)
        self.csv_export.setObjectName("GreenButton")

        self.excel_export = QPushButton("Export to Excel")
        self.excel_export.clicked.connect(self.export_to_excel)
        self.excel_export.setObjectName("GreenButton")
        
        self.delete_experiment_btn = QPushButton("ลบข้อมูลการทดลอง")
        self.delete_experiment_btn.clicked.connect(self.delete_selected_experiment)
        self.delete_experiment_btn.setObjectName("RedButton")

        self.csv_all_data = QPushButton("Export All Experiment to CSV")
        self.csv_all_data.clicked.connect(self.export_all_experiments_to_csv)
        self.csv_all_data.setObjectName("GreenButton")

        self.excel_all_data = QPushButton("Export All Experiment to Excel")
        self.excel_all_data.clicked.connect(self.export_all_experiments_to_excel)
        self.excel_all_data.setObjectName("GreenButton")

        self.theme_dropdown = QComboBox()
        self.theme_dropdown.setFixedSize(180, 32)
        self.theme_dropdown.setObjectName("ThemeDropdown")

        self.theme_dropdown.addItem("🌞 Light Theme", "light")
        self.theme_dropdown.addItem("🌜 Dark Theme", "dark")
        self.theme_dropdown.addItem("🌸 Pastel Theme", "pastel")
        self.theme_dropdown.addItem("🌈 Default Theme", "default")

        current_index = self.theme_dropdown.findData(self.main_window.get_theme())
        if current_index != -1:
            self.theme_dropdown.setCurrentIndex(current_index)

        self.theme_dropdown.currentIndexChanged.connect(self.change_theme)
        self.experiment_dropdown = self.create_experiment_dropdown()

        self.experiment_info = QLabel(f"ID ของการทดลอง: {self.current_experiment_id}\t\tชื่อการทดลอง: {self.current_experiment_name}\t\tวันที่ทดลอง: {self.current_experiment_date}\tประเภทการทดลอง: {self.current_experiment_type}")

        self.show_detail = QPushButton("Show Detail Note")
        self.show_detail.clicked.connect(self.show_detail_notes)
        self.show_detail.setObjectName("MainButton")

        self.edit_experiment = QPushButton("แก้ไขข้อมูลการทดลอง")
        self.edit_experiment.clicked.connect(self.edit_experiment_info)
        self.edit_experiment.setObjectName("YellowButton")

        self.bar_graph = FigureCanvas(plt.Figure(figsize=(4.5, 3.5)))
        self.line_graph = FigureCanvas(plt.Figure(figsize=(4.5, 3.5)))
        self.pie_graph = FigureCanvas(plt.Figure(figsize=(4.5, 3.5)))

        self.area_table = QTableWidget()
        self.area_table.setColumnCount(4)
        self.area_table.setHorizontalHeaderLabels(["ชื่อพื้นที่", "สีพื้นที่", "จำนวนการตรวจจับ/ครั้ง", "เวลาที่ตรวจจับได้/วินาที"])
        self.area_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.area_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.area_table.setStyleSheet("font-size: 14px;")

        self.main_layout = QVBoxLayout()

        top_row_layout = QHBoxLayout()
        top_row_layout.addWidget(self.experiment_info, stretch=2) 
        top_row_layout.addWidget(self.show_detail)              
        top_row_layout.addWidget(self.theme_dropdown)

        self.dropdown_layout = QVBoxLayout()
        self.dropdown_layout.addLayout(top_row_layout)
        self.dropdown_layout.addWidget(self.experiment_dropdown)

        self.card_frame = QFrame()
        self.card_frame.setFrameStyle(QFrame.Box | QFrame.Raised)
        self.card_frame.setLineWidth(2)
        self.card_layout = QGridLayout()
        self.card_frame.setLayout(self.card_layout)

        self.total_time_label = QLabel(f"ระยะเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.total_time}")
        self.total_time_label.setAlignment(Qt.AlignTop)
        self.total_time_label.setStyleSheet("font-size: 16px;")
        self.total_area_label = QLabel("จำนวนพื้นที่ที้งหมด:\n\n\n\n\t\t{self.total_area}")
        self.total_area_label.setAlignment(Qt.AlignTop)
        self.total_area_label.setStyleSheet("font-size: 16px;")
        self.experiment_time_label = QLabel(f"ระยะเวลาการทดลอง:\n\n\n\n\t\t{self.experiment_time}")
        self.experiment_time_label.setAlignment(Qt.AlignTop)
        self.experiment_time_label.setStyleSheet("font-size: 16px;")
        self.avg_time_label = QLabel(f"ค่าเฉลี่ยเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.avg_time}")
        self.avg_time_label.setAlignment(Qt.AlignTop)
        self.avg_time_label.setStyleSheet("font-size: 16px;")
        self.card_layout.addWidget(self.experiment_time_label, 0, 0)
        self.card_layout.addWidget(self.total_time_label, 0, 1)
        self.card_layout.addWidget(self.total_area_label, 1, 0)
        self.card_layout.addWidget(self.avg_time_label, 1, 1)
        self.card_layout.setContentsMargins(20, 20, 20, 20)

        self.graph_layout = QGridLayout()
        self.graph_layout.addWidget(self.card_frame, 0, 0)
        self.graph_layout.addWidget(self.bar_graph, 0, 1)
        self.graph_layout.addWidget(self.area_table, 1, 0)
        self.graph_layout.addWidget(self.pie_graph, 1, 1)
        self.graph_layout.setRowMinimumHeight(0, 350)
        self.graph_layout.setRowMinimumHeight(1, 350)
        self.graph_layout.setColumnMinimumWidth(0, 450)
        self.graph_layout.setColumnMinimumWidth(1, 450)
        self.graph_layout.setRowStretch(0, 1)
        self.graph_layout.setRowStretch(1, 1)
        self.graph_layout.setColumnStretch(0, 1)
        self.graph_layout.setColumnStretch(1, 1)
        self.main_layout.setStretchFactor(self.graph_layout, 1)   

        self.button_layout = QHBoxLayout()
        self.button_layout.addWidget(self.edit_experiment)
        self.button_layout.addWidget(self.delete_experiment_btn)
        self.button_layout.addWidget(self.csv_export)
        self.button_layout.addWidget(self.excel_export)
        self.button_layout.addWidget(self.csv_all_data)
        self.button_layout.addWidget(self.excel_all_data)
        self.button_layout.addStretch()
        self.button_layout.addWidget(self.switch_page)

        self.main_layout.addLayout(self.dropdown_layout) 
        self.main_layout.addLayout(self.graph_layout)                
        self.main_layout.addLayout(self.button_layout)    

        self.setLayout(self.main_layout)

        self.refresh()   

    def edit_experiment_info(self):
        preset_data = self.existing_data
        experiment_id = self.current_experiment_id

        if not experiment_id:
            QMessageBox.warning(self, "ยังไม่ได้เลือก", "กรุณาเลือกการทดลองก่อน")
            return

        dialog = ExperimentSetupDialog(self.main_window.db, self, preset_data=preset_data)
        if dialog.exec_() == QDialog.Accepted:
            experiment_type_id, name, date, detail, video_path = dialog.get_experiment_data()

            if experiment_type_id == 0:
                QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณาเลือกประเภทการทดลอง!")
                return

            success = self.main_window.db.update_experiment(
                experiment_id,
                experiment_type_id,
                name,
                date,
                detail if detail is not None else "ไม่มีรายละเอียดการทดลอง",
                video_path
            )

            if success:
                QMessageBox.information(self, "สำเร็จ", "แก้ไขข้อมูลการทดลองเรียบร้อยแล้ว")

                self.refresh()
                index = self.experiment_dropdown.findData(experiment_id)
                if index != -1:
                    self.experiment_dropdown.setCurrentIndex(index)
                    self.on_experiment_change(index)
            else:
                QMessageBox.warning(self, "ข้อผิดพลาด", "ไม่สามารถแก้ไขข้อมูลการทดลองได้")
    def show_detail_notes(self):
        experiment_id = self.current_experiment_id
        if not experiment_id:
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("รายละเอียดการทดลอง")
        dialog.resize(550, 400)  
        layout = QVBoxLayout(dialog)

        label = QLabel("สามารถแก้ไขรายละเอียดการทดลองได้")
        layout.addWidget(label)

        text_edit = QTextEdit()
        text_edit.setPlainText(self.current_experiment_note) 
        text_edit.setViewportMargins(10, 10, 10, 10)
        text_edit.setLineWrapMode(QTextEdit.NoWrap)  
        text_edit.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)  
        text_edit.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        layout.addWidget(text_edit)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec_() == QDialog.Accepted:
            new_detail = text_edit.toPlainText().strip()
            if new_detail:
                self.main_window.db.update_experiment_detail_note(experiment_id, new_detail)
                self.current_experiment_note = new_detail

    def update_theme_dropdown(self):
        current_theme = self.main_window.get_theme()
        index = self.theme_dropdown.findData(current_theme)
        if index != -1:
            self.theme_dropdown.blockSignals(True)
            self.theme_dropdown.setCurrentIndex(index)
            self.theme_dropdown.blockSignals(False)
            
    def delete_selected_experiment(self):
        experiment_id = self.current_experiment_id
        experiment_name = self.current_experiment_name

        if not experiment_id:
            QMessageBox.warning(self, "ยังไม่ได้เลือก", "กรุณาเลือกการทดลองก่อนลบ")
            return

        keyword = f"DELETE {experiment_id}"

        text, ok = QInputDialog.getText(
            self,
            "ยืนยันการลบ",
            "⚠️ การลบนี้ถาวรและกู้คืนไม่ได้\n\n"
            f"พิมพ์คำนี้เพื่อยืนยัน:\n{keyword}"
        )
        if not ok or text.strip() != keyword:
            QMessageBox.warning(self, "ยกเลิก", "ข้อความยืนยันไม่ถูกต้อง")
            return

        reply = QMessageBox.question(
            self,
            "ยืนยันอีกครั้ง",
            f"ต้องการลบการทดลองนี้จริงไหม?\n\nID: {experiment_id}\nชื่อ: {experiment_name}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        # ลบตามลำดับที่ถูกต้อง
        self.main_window.db.delete_raw_data_by_experiment_id(experiment_id)
        self.main_window.db.delete_area_summary_by_experiment_id(experiment_id)
        self.main_window.db.delete_experiment_by_id(experiment_id)

        QMessageBox.information(self, "สำเร็จ", "ลบข้อมูลการทดลองเรียบร้อยแล้ว")
        self.main_window.set_experiment_id(None)
        self.refresh()
    def load_graph_theme(self, theme):
        plt.rcdefaults()
        if theme == "dark":
            sns.set_theme(style="darkgrid")
            plt.rcParams.update({
                "figure.facecolor": "#121212",
                "axes.facecolor": "#1e1e1e",
                "axes.edgecolor": "#0f7f4f",
                "grid.color": "#2a4f4f",
                "axes.labelcolor": "#0f7f4f",
                "xtick.color": "#0f7f4f",
                "ytick.color": "#0f7f4f",
                "text.color": "#0f7f4f",
                "axes.titlecolor": "#0f7f4f",
                "legend.facecolor": "#1e1e1e",
                "legend.edgecolor": "#0f7f4f",
        })

        elif theme == "light":
            sns.set_theme(style="whitegrid")
            plt.rcParams.update({
                "figure.facecolor": "#ddecf9",      
                "axes.facecolor": "#ffffff",        
                "axes.edgecolor": "#336699",        
                "grid.color": "#cce0ff",            
                "axes.labelcolor": "#003366",       
                "xtick.color": "#003366",           
                "ytick.color": "#003366",           
                "text.color": "#003366",            
                "axes.titlecolor": "#00264d",       
                "legend.facecolor": "#e6f2ff",      
                "legend.edgecolor": "#336699",      
        })

        elif theme == "pastel":
            sns.set_theme(style="whitegrid", palette="pastel")
            plt.rcParams.update({
                "figure.facecolor": "#fdedf5",      
                "axes.facecolor": "#faf6f8",        
                "axes.edgecolor": "#d6a5c9",        
                "grid.color": "#f2d9e6",            
                "axes.labelcolor": "#cc6699",       
                "xtick.color": "#cc6699",           
                "ytick.color": "#cc6699",           
                "text.color": "#b34780",            
                "axes.titlecolor": "#b34780",       
                "legend.facecolor": "#ffe6f0",      
                "legend.edgecolor": "#ffcce7",      
        })

        elif theme == "default":
            sns.set_theme(style="whitegrid")
            plt.rcParams.update({
                "figure.facecolor": "#e6e6e6",      
                "axes.facecolor": "#f2f2f2",        
                "axes.edgecolor": "#999999",        
                "grid.color": "#cccccc",            
                "axes.labelcolor": "#444444",       
                "xtick.color": "#444444",           
                "ytick.color": "#444444",          
                "text.color": "#333333",           
                "axes.titlecolor": "#444444",       
                "legend.facecolor": "#e0e0e0",      
                "legend.edgecolor": "#b3b3b3",      
        })

    def change_theme(self, index):
        theme = self.theme_dropdown.itemData(index)
        if theme:
            self.main_window.set_theme(theme)
            self.main_window.load_theme(theme)

            self.load_graph_theme(theme)

            self.update_graph()
    
    def sanitize_filename(self,filename):
        invalid_chars = r'[<>:"|?*]'
        sanitized = re.sub(invalid_chars, '_', filename)
        sanitized = sanitized.strip().strip('.')
        return sanitized
    
    def _write_csv(self, data, output_file, headers):
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                for row in data:
                    writer.writerow(row)
        except OSError as e:
            print(f"Error writing to CSV file {output_file}: {e}")
            raise

    def _write_excel(self, data, worksheet, headers):
        if not isinstance(worksheet, Worksheet):
            raise TypeError(f"Expected Worksheet object, got {type(worksheet).__name__} instead")
        
        try:
            worksheet.append(headers)
            for row in data:
                worksheet.append(row)
        except Exception as e:
            print(f"Error writing to Excel worksheet {worksheet.title}: {e}")
            raise
    
    def export_to_csv(self):
        if not self.area_summary and not self.raw_data:
            print("Not found area summary and rawdata")
            return

        out_dir = QFileDialog.getExistingDirectory(self, "เลือกโฟลเดอร์สำหรับบันทึก CSV", "")
        if not out_dir:
            return 
        
        date = str(self.current_experiment_date)
        name = str(self.current_experiment_name)
        safe_base = self.sanitize_filename(f"{date}_{name}")

        if self.current_experiment_id:
            experiment_info = self.main_window.db.get_experiment_by_id(self.current_experiment_id)

            experiment_file = os.path.join(out_dir, f"{safe_base}_experiment_info.csv")
            experiment_headers = ['experiment_id', 'experiment_type_id', 'experiment_name', 'experiment_date', 'experiment_note', 'video_path']
            self._write_csv([experiment_info], experiment_file, experiment_headers)

            experiment_type = self.main_window.db.get_experiment_types()
            experiment_type_file = os.path.join(out_dir, f"{safe_base}_experiment_type.csv")
            experiment_type_headers = ['experiment_type_id', 'type_name']
            self._write_csv(experiment_type, experiment_type_file, experiment_type_headers)

        if self.area_summary:
            summary_file = os.path.join(out_dir, f"{safe_base}_area_summary.csv")
            summary_headers = ['area_id', 'experiment_id', 'area_name', 'color', 'hit_count', 'total_time', 'area_point']
            self._write_csv(self.area_summary, summary_file, summary_headers)

        if self.raw_data:
            rawdata_file = os.path.join(out_dir, f"{safe_base}_raw_data.csv")
            rawdata_headers = ['experiment_id', 'area_id', 'timestamp', 'frame_count', 'area_name', 'rat_position_x', 'rat_position_y']
            self._write_csv(self.raw_data, rawdata_file, rawdata_headers)

        QMessageBox.information(self, "สำเร็จ", f"บันทึกไฟล์ CSV แล้วที่:\n{out_dir}")

    def export_to_excel(self):
        if not self.area_summary and not self.raw_data:
            print("No data to export: both area_summary and raw_data are empty")
            return

        date = str(self.current_experiment_date)
        name = str(self.current_experiment_name)
        safe_base = self.sanitize_filename(f"{date}_{name}_data")

        default_name = f"{safe_base}.xlsx"

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "บันทึกไฟล์ Excel",
            default_name,
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            return  

        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        output_dir = os.path.dirname(save_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        try:
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            if self.current_experiment_id:
                experiment_info = self.main_window.db.get_experiment_by_id(self.current_experiment_id)
                headers = ['experiment_id', 'experiment_type_id', 'experiment_name', 'experiment_date', 'experiment_note', 'video_path']
                ws_experiment_info = workbook.create_sheet(title="Experiment Info")
                self._write_excel([experiment_info], ws_experiment_info, headers)

                experiment_type = self.main_window.db.get_experiment_types()
                headers = ['experiment_type_id', 'type_name']
                ws_experiment_type = workbook.create_sheet(title="Experiment Type")
                self._write_excel(experiment_type, ws_experiment_type, headers)

            if self.area_summary:
                headers = ['area_id', 'experiment_id', 'area_name', 'color', 'hit_count', 'total_time', 'area_point']
                ws_summary = workbook.create_sheet(title="Area Summary")
                self._write_excel(self.area_summary, ws_summary, headers)

            if self.raw_data:
                headers = ['experiment_id', 'area_id', 'time_stamp', 'frame_count', 'area_name', 'rat_position_x', 'rat_position_y']
                ws_raw_data = workbook.create_sheet(title="Raw Data")
                self._write_excel(self.raw_data, ws_raw_data, headers)

            workbook.save(save_path)
            QMessageBox.information(self, "สำเร็จ", f"บันทึกไฟล์ Excel แล้วที่:\n{save_path}")

        except Exception as e:
            QMessageBox.critical(self, "ข้อผิดพลาด", f"Export ล้มเหลว:\n{e}")
            raise
    def export_all_experiments_to_csv(self):

        out_dir = QFileDialog.getExistingDirectory(
            self,
            "เลือกโฟลเดอร์สำหรับบันทึก CSV (ทุกการทดลอง)",
            ""
        )
        if not out_dir:
            return  

        try:
            all_experiments = self.main_window.db.get_all_experiments()
            experiments_file = os.path.join(out_dir, "all_experiments.csv")
            self._write_csv(
                all_experiments,
                experiments_file,
                ['experiment_id', 'experiment_type_id', 'experiment_name', 'experiment_date', 'experiment_note', 'video_path']
            )

            all_experiment_types = self.main_window.db.get_experiment_types()
            types_file = os.path.join(out_dir, "all_experiment_types.csv")
            self._write_csv(
                all_experiment_types,
                types_file,
                ['experiment_type_id', 'type_name']
            )

            all_area_summary = self.main_window.db.get_all_area_summary()
            area_file = os.path.join(out_dir, "all_area_summary.csv")
            self._write_csv(
                all_area_summary,
                area_file,
                ['area_id', 'experiment_id', 'area_name', 'color', 'hit_count', 'total_time', 'area_point']
            )

            all_raw_data = self.main_window.db.get_all_raw_data()
            raw_file = os.path.join(out_dir, "all_raw_data.csv")
            self._write_csv(
                all_raw_data,
                raw_file,
                ['experiment_id', 'area_id', 'timestamp', 'frame_count', 'area_name', 'rat_position_x', 'rat_position_y']
            )

            QMessageBox.information(
                self,
                "Export สำเร็จ",
                f"บันทึก CSV ทุกการทดลองเรียบร้อยแล้ว\nที่: {out_dir}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Export ล้มเหลว", str(e))

    def export_all_experiments_to_excel(self):
        default_name = "all_experiments_data.xlsx"

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "บันทึกไฟล์ Excel (ทุกการทดลอง)",
            default_name,
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            return 

        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        try:
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            all_experiments = self.main_window.db.get_all_experiments()
            ws_exp = workbook.create_sheet("All Experiments")
            self._write_excel(
                all_experiments,
                ws_exp,
                ['experiment_id', 'experiment_type_id', 'experiment_name', 'experiment_date', 'experiment_note', 'video_path']
            )

            all_experiment_types = self.main_window.db.get_experiment_types()
            ws_type = workbook.create_sheet("All Experiment Types")
            self._write_excel(
                all_experiment_types,
                ws_type,
                ['experiment_type_id', 'type_name']
            )

            all_area_summary = self.main_window.db.get_all_area_summary()
            ws_area = workbook.create_sheet("All Area Summary")
            self._write_excel(
                all_area_summary,
                ws_area,
                ['area_id', 'experiment_id', 'area_name', 'color', 'hit_count', 'total_time', 'area_point']
            )

            all_raw_data = self.main_window.db.get_all_raw_data()
            ws_raw = workbook.create_sheet("All Raw Data")
            self._write_excel(
                all_raw_data,
                ws_raw,
                ['experiment_id', 'area_id', 'timestamp', 'frame_count', 'area_name', 'rat_position_x', 'rat_position_y']
            )

            workbook.save(save_path)

            QMessageBox.information(
                self,
                "Export สำเร็จ",
                f"บันทึกไฟล์ Excel ทุกการทดลองเรียบร้อยแล้ว\nที่:\n{save_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Export ล้มเหลว", str(e))

        
    def create_experiment_dropdown(self):
        dropdown = QComboBox()
        all_experiments = self.main_window.db.get_all_experiments()
        dropdown.addItem("เลือกการทดลอง", 0)
        for experiment in all_experiments:
            dropdown.addItem(experiment[2], experiment[0])
        
        dropdown.setFixedHeight(40)
        dropdown.setFont(QFont("Arial", 20))
        dropdown.currentIndexChanged.connect(self.on_experiment_change)

        return dropdown

    def update_table(self):
        if self.df_area_summary is None or self.df_area_summary.empty:
            self.area_table.setRowCount(0)
            self.area_table.setRowCount(1)
            item = QTableWidgetItem("ไม่มีข้อมูลพื้นที่")
            item.setTextAlignment(Qt.AlignCenter)
            self.area_table.setItem(0, 0, item)
            self.area_table.setSpan(0, 0, 1, self.area_table.columnCount())
            return

        self.area_table.clearSpans()

        df = self.df_area_summary.copy()

        df = df.drop_duplicates(subset=["Area Name"], keep="last").reset_index(drop=True)

        self.area_table.setRowCount(len(df))

        for row, (_, data) in enumerate(df.iterrows()):
            color_hex = str(data["Color"])

            values = [
                str(data["Area Name"]),
                "",  
                str(data["Hit Count"]) + "  ครั้ง",
                str(data["Total Time"]) + "  วินาที",
            ]

            for col, val in enumerate(values):
                item = QTableWidgetItem(val)

                if col == 1:
                    item.setBackground(QColor(color_hex))
                    item.setText("")

                self.area_table.setItem(row, col, item)
    def update_graph(self):
        self.bar_graph.figure.clear()
        self.line_graph.figure.clear()
        self.pie_graph.figure.clear()

        has_area_summary = self.df_area_summary is not None and not self.df_area_summary.empty
        has_raw_data = self.df_raw_data is not None and not self.df_raw_data.empty

        try:
            ax_line = self.line_graph.figure.add_subplot(111)
            if has_raw_data:
                sns.lineplot(data=self.df_raw_data, x="Timestamp", y="X", hue="Area Name", ax=ax_line)
                ax_line.set_title("x position per time")
                ax_line.set_xlabel("Time (seconds)")
                ax_line.set_ylabel("X position")
            else:
                ax_line.text(0.5, 0.5, "No data available\nOr invalid data", ha='center', va='center', fontsize=14)
                ax_line.set_axis_off()
            self.line_graph.figure.tight_layout()

            ax_bar = self.bar_graph.figure.add_subplot(111)
            ax_pie = self.pie_graph.figure.add_subplot(111)
            if has_area_summary:
                sns.barplot(data=self.df_area_summary, x="Area Name", y="Hit Count", hue="Area Name", palette=list(self.df_area_summary["Color"]), legend=False, ax=ax_bar, edgecolor='#444444', linewidth=1.5)
                for i, v in enumerate(self.df_area_summary["Hit Count"]):
                    ax_bar.text(i, (v/2), str(v), ha='center', va='bottom', fontsize=12)
                ax_bar.set_title("Area per number of detections")
                ax_bar.set_xlabel("Area Name")
                ax_bar.set_ylabel("Number of Detections")

                if self.df_area_summary["Total Time"].sum() > 0:
                    ax_pie.pie(self.df_area_summary["Total Time"], labels=self.df_area_summary["Area Name"], colors=self.df_area_summary["Color"], autopct=lambda pct: f"{pct:.1f}%\n({pct * self.df_area_summary['Total Time'].sum() / 100:.1f} seconds)", wedgeprops={'edgecolor': 'black', 'linewidth': 1.5})
                    ax_pie.set_title("Area per total time")
                    ax_pie.axis('equal')
                else : 
                    print("No valid data available for pie graph")
                    ax_pie.text(0.5, 0.5, "No data available\nOr invalid data", ha='center', va='center', fontsize=14)
                    ax_pie.set_axis_off()
            else:
                # print("No valid data available for bar and pie graphs")
                ax_bar.text(0.5, 0.5, "No data available\nOr invalid data", ha='center', va='center', fontsize=14)
                ax_pie.text(0.5, 0.5, "No data available\nOr invalid data", ha='center', va='center', fontsize=14)
                ax_bar.set_axis_off()
                ax_pie.set_axis_off()
            
            self.pie_graph.figure.tight_layout()
            self.bar_graph.figure.tight_layout()


        except Exception as e:
            print(f"Error generating graphs: {e}")
            ax_bar.text(0.5, 0.5, f"Error occurred: {str(e)}", ha='center', va='center', fontsize=14)
            ax_pie.text(0.5, 0.5, f"Error occurred: {str(e)}", ha='center', va='center', fontsize=14)
            ax_line.text(0.5, 0.5, f"Error occurred: {str(e)}", ha='center', va='center', fontsize=14)
            ax_bar.set_axis_off()
            ax_pie.set_axis_off()
            ax_line.set_axis_off()
            self.bar_graph.figure.tight_layout()
            self.pie_graph.figure.tight_layout()
            self.line_graph.figure.tight_layout()

        self.bar_graph.draw()
        self.line_graph.draw()
        self.pie_graph.draw()

    def prepare_data_for_graph(self):
        self.df_raw_data = None
        self.df_area_summary = None
        self.experiment_time = "ไม่มีปรากฏข้อมูล"
        self.total_area = "ไม่มีปรากฏข้อมูล"
        self.total_time = "ไม่มีปรากฏข้อมูล"
        self.avg_time = "ไม่มีปรากฏข้อมูล"

        def parse_color(color_str):
            try:
                b, g, r = map(int, color_str.strip('()').split(','))
                return f'#{r:02x}{g:02x}{b:02x}'
            except:
                return color_str
            
        if self.raw_data:
            try:
                self.df_raw_data = pd.DataFrame(self.raw_data, columns=["Experiment ID", "Area ID", "Timestamp", "Frame Count", "Area Name", "X", "Y"])
                self.df_raw_data["Timestamp"] = pd.to_numeric(self.df_raw_data["Timestamp"], errors="coerce")
                self.df_raw_data["X"] = pd.to_numeric(self.df_raw_data["X"], errors="coerce")
                self.df_raw_data["Y"] = pd.to_numeric(self.df_raw_data["Y"], errors="coerce")
                self.df_raw_data["Area Name"] = self.df_raw_data["Area Name"].fillna("Unknown")

                self.experiment_time = str(self.df_raw_data["Timestamp"].max()) + "  วินาที"
            except Exception as e:
                print(f"Error preparing raw_data: {e}")

        if self.area_summary:
            try:
                self.df_area_summary = pd.DataFrame(self.area_summary, columns=["ID", "Experiment ID", "Area Name", "Color", "Hit Count", "Total Time", "Area Point"])
                self.df_area_summary["Hit Count"] = pd.to_numeric(self.df_area_summary["Hit Count"], errors="coerce")
                self.df_area_summary["Total Time"] = pd.to_numeric(self.df_area_summary["Total Time"], errors="coerce")
                self.df_area_summary["Color"] = self.df_area_summary["Color"].apply(parse_color)

                self.total_time = str(self.df_area_summary["Total Time"].sum()) + "  วินาที"
                self.total_area = str(len(self.df_area_summary)) + "  พื้นที่"
                self.avg_time = str(round(self.df_area_summary["Total Time"].mean(), 2)) + "  วินาที"
            except Exception as e:
                print(f"Error preparing area_summary: {e}")

        self.experiment_time_label.setText(f"ระยะเวลาการทดลอง:\n\n\n\n\t\t{self.experiment_time}")
        self.total_area_label.setText(f"จำนวนพื้นที่ที้งหมด:\n\n\n\n\t\t{self.total_area}")
        self.total_time_label.setText(f"ระยะเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.total_time}")
        self.avg_time_label.setText(f"ค่าเฉลี่ยเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.avg_time}")
        
        self.update_graph()
        self.update_table()

    def refresh(self):
        self.experiment_dropdown.blockSignals(True)
        self.experiment_dropdown.clear()
        all_experiments = self.main_window.db.get_all_experiments()
        self.experiment_dropdown.addItem("เลือกการทดลอง", 0)
        for experiment in all_experiments:
            self.experiment_dropdown.addItem(experiment[2], experiment[0])
        self.experiment_dropdown.blockSignals(False)

        experiment_id = self.main_window.get_experiment_id()
        index = self.experiment_dropdown.findData(experiment_id) if experiment_id else 0
        self.experiment_dropdown.setCurrentIndex(index)
        self.on_experiment_change(index)
    
    def on_experiment_change(self, index):
        self.area_summary = None
        self.raw_data = None

        if index > 0:
            selected_id = self.experiment_dropdown.itemData(index)
            self.area_summary = self.main_window.db.get_area_summary_by_experiment_id(selected_id)
            self.raw_data = self.main_window.db.get_raw_data_by_experiment_id(selected_id)
            self.current_experiment_id = selected_id
            experiment = self.main_window.db.get_experiment_by_id(selected_id)
            self.current_experiment_name = experiment[2] if experiment else "None"
            self.current_experiment_date = experiment[3] if experiment else "None"
            self.current_experiment_note = experiment[4] if experiment else "None"
            self.current_experiment_type = self.main_window.db.get_experiment_type_by_id(experiment[1])
            self.experiment_info.setText(f"ID ของการทดลอง: {self.current_experiment_id}\t\tชื่อการทดลอง: {self.current_experiment_name}\t\tวันที่ทดลอง: {self.current_experiment_date}\tประเภทการทดลอง: {self.current_experiment_type}")
            self.existing_data = {
                "name" : experiment[2],
                "type_id" : experiment[1],
                "detail" : experiment[4],
                "date" : experiment[3],
                "video_path" : experiment[5]
            }

            # if self.area_summary and self.raw_data:
            #     print(f"\nLoad Data success Experiment ID: {self.current_experiment_id} Experiment Name: {self.current_experiment_name}\narea_summary:\n{self.area_summary[0]}\nraw_data:\n{self.raw_data[0]}")
            # elif self.area_summary or self.raw_data:
            #     print(f"\nLoad Data incomplete Experiment ID: {self.current_experiment_id} Experiment Name: {self.current_experiment_name}\narea_summary:\n{self.area_summary[0] if self.area_summary else None}\nraw_data:\n{self.raw_data[0] if self.raw_data else None}")
            # else:
            #     print(f"\nLoad Data fail Experiment ID: {self.current_experiment_id} Experiment Name: {self.current_experiment_name}")
            #     print(f"Area summary data not found: {self.area_summary}")
            #     print(f"Raw data not found: {self.raw_data}")
        else:
            print("No experiment selected")
            self.clear_data()

        self.prepare_data_for_graph()
    
    def clear_data(self):
        self.area_summary = None
        self.raw_data = None
        self.df_raw_data = None
        self.df_area_summary = None
        self.current_experiment_id = None
        self.current_experiment_name = None
        self.current_experiment_date = None
        self.current_experiment_type = None
        self.current_experiment_note = None
        self.existing_data = None
        self.total_time = "ไม่มีปรากฏข้อมูล"
        self.total_area = "ไม่มีปรากฏข้อมูล"
        self.experiment_time = "ไม่มีปรากฏข้อมูล"
        self.avg_time = "ไม่มีปรากฏข้อมูล"
        self.experiment_time_label.setText(f"ระยะเวลาการทดลอง:\n\n\n\n\t\t{self.experiment_time}")
        self.total_area_label.setText(f"จำนวนพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.total_area}")
        self.total_time_label.setText(f"ระยะเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.total_time}")
        self.experiment_info.setText(f"ID ของการทดลอง: {self.current_experiment_id}\t\tชื่อการทดลอง: {self.current_experiment_name}\t\tวันที่ทดลอง: {self.current_experiment_date}\tประเภทการทดลอง: {self.current_experiment_type}")
        self.avg_time_label.setText(f"ค่าเฉลี่ยเวลาของพื้นที่ทั้งหมด:\n\n\n\n\t\t{self.avg_time}")

    def switch_to_home_page(self):
        self.experiment_dropdown.setCurrentIndex(self.experiment_dropdown.findData(0))
        self.main_window.set_experiment_id(None)
        self.clear_data()
        self.main_window.switch_to_page(0)
        

        
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_theme = "default"
        self.setWindowTitle("mice detection")
        icon_path = resource_path("assets", "mouse.png")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.fps = 30

        self.stack = QStackedWidget()
        self.video_path = None
        self.experiment_id = None
        self.experiment_name = None

        self.db = DatabaseManager()

        self.page1 = PageOne(self.stack, self)
        self.page2 = PageTwo(self.stack, self)
        self.page3 = PageThree(self.stack, self)

        self.stack.addWidget(self.page1)
        self.stack.addWidget(self.page2)
        self.stack.addWidget(self.page3)

        self.setCentralWidget(self.stack)
        self.setCentralWidget(self.stack)

        self.setFixedSize(self.page1.size())
        self.center_window()

        self.apply_standard_ui_colors()   

        self.load_theme(self.current_theme)

        self.setFixedSize(self.page1.size()) 

        self.center_window()  

        self.load_theme(self.current_theme)

    def switch_to_page(self, index):
        self.stack.setCurrentIndex(index)
        current_widget = self.stack.currentWidget()
        self.setFixedSize(current_widget.size()) 
        self.center_window()  

        if index == 0 and isinstance(current_widget, PageOne):
            current_widget.update_theme_dropdown()

        if index == 1 and isinstance(current_widget, PageTwo):
            current_widget.update_video()

        if index == 2 and isinstance(current_widget, PageThree):  
            current_widget.refresh()
            current_widget.update_theme_dropdown()

    def center_window(self):
        window_rect = self.frameGeometry()
        center_point = QDesktopWidget().availableGeometry().center()
        window_rect.moveCenter(center_point)
        self.move(window_rect.topLeft()) 
    def apply_standard_ui_colors(self):
        self.setStyleSheet("""
            QPushButton {
                border: 1px solid rgba(0,0,0,0.18);
                border-radius: 10px;
                padding: 8px 14px;
                font-weight: 600;
            }

            QPushButton:hover {
                border: 1px solid rgba(0,0,0,0.28);
            }

            QPushButton:pressed {
                border: 1px solid rgba(0,0,0,0.35);
            }

            QPushButton#MainButton {
                background-color: #2563EB;   
                color: #FFFFFF;              
            }

            QPushButton#GreenButton {
                background-color: #16A34A;  
                color: #FFFFFF;              
            }
            QPushButton#RedButton {
                background-color: #DC2626;   
                color: #FFFFFF;              
            }

            QPushButton#YellowButton {
                background-color: #F59E0B;   
                color: #1F2937;             
            }

            QLineEdit, QTextEdit, QComboBox {
                color: #111827;
            }
        """)
        
    def load_theme(self, theme_name):
        theme_files = {
            "light": resource_path("style", "light.qss"),
            "dark": resource_path("style", "dark.qss"),
            "default": resource_path("style", "default.qss"),
            "pastel": resource_path("style", "pastel.qss")
        }
        if theme_name not in theme_files:
            print(f"Error: Theme '{theme_name}' not supported.")
            return
        try:
            with open(theme_files[theme_name], "r", encoding="utf-8") as style_file:
                app = QApplication.instance()
                app.setStyleSheet(style_file.read())
                self.current_theme = theme_name
        except FileNotFoundError:
            print(f"Error: QSS file '{theme_files[theme_name]}' not found.")
        except Exception as e:
            print(f"Error loading theme '{theme_name}': {e}")

    def set_video_path(self, video_path):
        self.video_path = video_path

    def get_video_path(self):
        return self.video_path
    
    def set_fps(self, fps):
        self.fps = fps

    def get_fps(self):
        return self.fps
    
    def set_experiment_name(self, experiment_name):
        self.experiment_name = experiment_name

    def get_experiment_name(self):
        return self.experiment_name
    
    def set_experiment_id(self, experiment_id):
        self.experiment_id = experiment_id

    def get_experiment_id(self):
        return self.experiment_id

    def set_theme(self, theme_name):
        self.current_theme = theme_name

    def get_theme(self):
        return self.current_theme



def main():
    try:
        app = QApplication(sys.argv)
        window = MainWindow()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Fatal error: {e}")
        raise

if __name__ == "__main__":
    main()